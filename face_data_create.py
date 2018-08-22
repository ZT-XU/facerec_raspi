from PIL import Image
import face_recognition
import picamera
import numpy as np
import openpyxl


wb_name = ''
res = input('Do you want to create a new excel file?(yes/no)')
if 'yes' == res.lower():
    wb = openpyxl.Workbook()
    ws = wb.active
    num = 0
else:
    wb_name = input('Please Enter the name of the excel file:')
    wb = openpyxl.load_workbook(wb_name + '.xlsx')
    ws = wb.active
    num = 1

value_list = []

ID = input("Enter your name:")
for cells in ws['A']:
    value_list.append(cells.value)
    num = num + 1
if ID not in value_list:
    cell = ws.cell(row=num, column=1, value=ID)

camera = picamera.PiCamera()
camera.resolution = (320, 240)
output = np.empty((240, 320, 3), dtype=np.uint8)

while(ID):
    print("Capturing image.")
    # Grab a single frame of video from the RPi camera as a numpy array
    camera.capture(output, format="rgb")
    face_locations = face_recognition.face_locations(output, number_of_times_to_upsample=0, model="cnn")


    for face_location in face_locations:

        # Print the location of each face in this image
        top, right, bottom, left = face_location
        

        # You can access the actual face itself like this:
        face_image = image[top:bottom, left:right]
        pil_image = Image.fromarray(face_image)
        g_pil_image = pil_image.convert('L')
        with open('./images' + ID + '.png', 'wb') as f:
            f.write(g_pil_image)
        print('Image Get!!!')
        ID = False
if not wb_name:
    wb_name = input('Please Enter the name of the excel file:')
wb.save('./workbooks/' + wb_name + '.xlsx')
print("Finish!!!")
