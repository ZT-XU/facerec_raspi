from PyQt5 import QtGui
import sys
from subprocess import Popen, PIPE
from PyQt5.QtWidgets import QApplication, QDialog, QVBoxLayout,\
     QGroupBox, QPushButton, QTextBrowser, QLabel, QGridLayout,  QMenuBar,QAction,QInputDialog,\
    QMessageBox, QProgressBar, QFileDialog
import SSHConnection
from Dialog import Dialog, OneLineDialog
from PyQt5.QtCore import QTime, QTimer, QDateTime, Qt, QBasicTimer, QUrl
from PyQt5.QtGui import QDesktopServices
import os
from Table import Table, TaskTable
import openpyxl


class Window(QDialog):
    def __init__(self):
        super().__init__()
        # 创建定时器对象和时间对象
        self.isTime = False
        self.timer = QTimer()
        self.timer2 = QBasicTimer()
        self.timeStep = 0
        self.step = 0
        self.timeClock = QTime()
        self.time_label = QLabel('剩余：')
        self.label_time_val = QLabel(self)
        self.date = QLabel('日期:')
        self.date_time = QDateTime.currentDateTime()
        self.date_val = QLabel(self.date_time.toString())
        self.timer.start(1000)  # 启动定时器，定时器对象每隔一秒发射一个timeout信号
        self.num = 0
        self.processbar = QProgressBar(self)
        self.textbrowser = QTextBrowser()
        self.ISCONNECTED = False
        self.CANCEL = False
        self.excel_file_name = ""
        # Start button
        self.btn4 = QPushButton('开始签到')
        self.conn = None
        self.username = ""
        self.hostname = ""
        self.password = ""
        self.PRINT_LOG = True
        self.task_name = ""
        self.now_log = ""


        self.title = "人脸识别签到系统"
        self.top = 400
        self.left = 200
        self.width = 1024
        self.height = 720
        self.log = ""
        self.a = False
        self.b = False
        self.p1 = None
        self.p2 = None
        self.p3 = None
        self.Qfile = None
        self.id_list = []
        self.name_list = []
        self.temp = ""

        self.InitWindow()


    def InitWindow(self):
        self.setWindowTitle(self.title)
        self.setWindowIcon(QtGui.QIcon("icon.png"))
        self.setGeometry(self.top, self.left, self.width, self.height)


        # 全局布局
        v1 = QVBoxLayout()
        grid = QGridLayout()
        group_box = QGroupBox()

        menu = QMenuBar(self)
        new_menu = menu.addMenu('新的任务')
        create_menu = menu.addMenu('创建faceID')
        connect_menu = menu.addMenu('连接到...')
        result_menu = menu.addMenu('查看')

        # Action
        load_button = QAction("从名单载入签到任务",self);
        new_menu.addAction(load_button)
        new_button = QAction('手动创建一个新的签到任务',self)
        new_menu.addAction(new_button)
        view_button = QAction('载入已存在的签到任务',self)
        new_menu.addAction(view_button)
        create_button = QAction('创建faceID', self)
        create_menu.addAction(create_button)
        connect_button = QAction('连接到...', self)
        connect_menu.addAction(connect_button)
        result_button_now = QAction('查看当前签到结果',self)
        result_menu.addAction(result_button_now)
        result_button_before = QAction('查看历史签到结果',self)
        result_menu.addAction(result_button_before)

        # connect function
        #connect_button.triggered.connect(self.Isconnected) delete
        connect_button.triggered.connect(self.connect_function)

        # 新建签到任务
        new_button.triggered.connect(self.new_button_function)

        # 载入签到任务
        load_button.triggered.connect(self.load_task)

        # 查看已有的签到任务
        view_button.triggered.connect(self.view_button_function)

        # 创建faceID
        create_button.triggered.connect(self.create_button_function)

        # 查看结果
        result_button_now.triggered.connect(self.result_button_now_function)
        result_button_before.triggered.connect(self.result_button_before_function)

        exit_button = QPushButton('退出')
        exit_button.clicked.connect(self.close)
        grid.addWidget(self.time_label,1,0)
        grid.addWidget(self.label_time_val,1,1)
        grid.addWidget(self.date,0,0)
        grid.addWidget(self.date_val,0,1)
        group_box.setLayout(grid)

        # start button function
        self.btn4.clicked.connect(self.settime)


        self.timer.timeout.connect(self.showdate)

        v1.addWidget(menu)
        v1.addWidget(self.btn4)

        v1.addWidget(group_box)
        v1.addWidget(self.processbar)
        v1.addWidget(self.textbrowser)
        v1.addWidget(exit_button)

        self.process = True  # 来控制识别窗口的进行
        with open("process.txt", 'w') as f:
            f.write(str(self.process))

        with open("proceed.txt", 'w') as f:
            f.write("False")
        self.setLayout(v1)

        self.show()

    def addtime(self):  # 计时时间增一秒，并显示在QLable上
        if self.num > 0:
            self.label_time_val.setText(str(self.num // 60) + ' 分钟' + str(self.num % 60) + ' 秒')  # 标签显示时间
            self.num -= 1

        if self.num == 0 and self.isTime:
            self.label_time_val.setText('0 分钟')
            self.endMessage()
            self.textbrowser.append("签到结束!!!")
            self.process = not self.process
            with open("process.txt", 'w') as f:
                f.write(str(self.process))
            os.system("ps a | grep exe_now.py > /home/xzt/facetest/Window/kill.txt")
            os.system("ps a | grep exe_now.sh > /home/xzt/facetest/Window/kill2.txt")
            with open("/home/xzt/facetest/Window/kill.txt", "r") as f:
                r1 = f.read(5)
            with open("/home/xzt/facetest/Window/kill2.txt", "r") as f:
                r2 = f.read(5)
            os.system("kill " + r2)
            os.system("kill "+r1)
            self.btn4.setEnabled(True)
            os.remove("/home/xzt/facetest/Window/now_log.txt")
            self.conn.download("/home/pi/Project/workbooks/" + self.task_name + '.xlsx', "/home/xzt/facetest/result/" + self.task_name + '.xlsx')

    def showdate(self):
        datetime = QDateTime.currentDateTime()
        text = datetime.toString()
        self.date_val.setText(text)

    def settime(self):
        if self.task_name:
            text, self.isTime = QInputDialog.getInt(self, '设置签到时间', '请输入签到时间(分钟)：', min=1)
            if self.isTime:
                if self.timeStep != 0:
                    self.timeStep = 0
                if self.num != 0:
                    self.num = 0
                self.num = int(text) * 60
                self.timeStep = self.num * 10
                #self.onStart()
                self.p1 = Popen("/home/xzt/facetest/exe_now.sh", stdout=PIPE, close_fds=True)
                #self.p2 = Popen("/home/xzt/facetest/exe_now_before.sh", stdout=PIPE, close_fds=True)
                #self.p3 = Popen("/home/xzt/facetest/refresh.py", stdout=PIPE, close_fds=True)
                self.textbrowser.append("准备开始签到...")
                self.textbrowser.append("程序启动需要时间,请稍等...")
                self.btn4.setEnabled(False)
                if self.num > 0:
                    #self.timer.timeout.connect(self.addtime)
                    self.timer.timeout.connect(self.get_now_log)


        else:
            QMessageBox.about(self,"错误","没有检测到任务,请先载入任务")

    def endMessage(self):
        QMessageBox.about(self, "签到结束", "签到已经结束,您现在可以查看签到结果")

        self.isTime = False

    def connect_function(self):
        dialog = Dialog()
        dialog.exec_()
        self.username = dialog.getUsername()
        self.hostname = dialog.getHostname()
        self.password = dialog.getPassword()
        self.CANCEL = dialog.CANCEL
        if not self.CANCEL:
            self.ssh(self.username,self.hostname,self.password)
            self.conn.download('/home/pi/Project/FACEID/Face-ID.xlsx','/home/xzt/facetest/FACEID/Face-ID.xlsx')
            #self.conn.put('/home/xzt/facetest/Window/train/trained_knn_model.clf',"/home/pi/Project/trained_knn_model.clf")

    def Isconnected(self):
        self.ISCONNECTED = True

    #def set_excel_file_name(self):
        #file_dialog = OneLineDialog("请输入签到任务名","任务名:")
        #file_dialog.exec_()
        #self.excel_file_name = file_dialog.text

    def ssh(self, username, hostname, password):
        self.conn = SSHConnection.SSHConnection(username, hostname, password)
        self.Isconnected()
        info = self.conn.exec_command('ls -al')
        if len(info) > 0:
            self.textbrowser.append("Successful Connection!!!")

    def timerEvent(self, event):
        if self.step >= 100:
            self.timer2.stop()
            self.step = 0
            return
        self.step = self.step + 1
        self.processbar.setValue(self.step)

    def onStart(self):
        if self.timer2.isActive():
            self.timer2.stop()
        else:
            self.timer2.start(self.timeStep,self)

    # 新建任务 && 创建一个新的签到任务
    def new_button_function(self):
        table = Table()
        table.exec_()

    # 载入一个签到任务
    def load_task(self):
        if self.ISCONNECTED:
            self.Qfile = QFileDialog()
            fname = self.Qfile.getOpenFileName(self, '选择文件', '/home/xzt/')[0]
            print(fname)
            if 'xlsx' not in fname:
                QMessageBox.about("错误","不支持的文件格式")
            else:
                wb = openpyxl.load_workbook(fname)
                ws = wb.active
                count = 0
                for cell in ws['A']:
                    value = cell.value
                    count += 1
                    if value is None:
                        break
                    if value == '学号':
                        continue
                    self.id_list.append(value)
                    self.name_list.append(ws['B' + str(count)].value)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws['A1'] = "学号"
                ws['B1'] = "姓名"
                ws["C1"] = "是否签到"
                ws["D1"] = "签到时间"
                ws.merge_cells("D1:F1")
                for i in range(count-2):
                    ws['A' + str(i + 2)] = self.id_list[i]
                    ws["B" + str(i + 2)] = self.name_list[i]
                dialog = OneLineDialog("请输入新建任务名", "任务名")
                dialog.exec_()
                task_name = dialog.text
                cancel = dialog.CANCEL
                if not cancel:
                    wb.save("/home/xzt/facetest/workbooks/" + task_name + '.xlsx')
                    self.task_name = task_name
                    QMessageBox.about(self,"成功","成功创建任务：" + "\n" + task_name)
                    with open('task_name.txt', 'a') as f:
                        f.write(task_name + '\n')
                    self.conn.put("/home/xzt/facetest/workbooks/" + self.task_name + ".xlsx",
                                  "/home/pi/Project/workbooks/" + self.task_name + ".xlsx")
                    with open("now_task_name.txt", 'w', encoding="gbk") as f:
                        f.write(self.task_name)
                    self.conn.put("/home/xzt/facetest/Window/now_task_name.txt", "/home/pi/Project/now_task_name.txt")
                    print(self.task_name)
        else:
            QMessageBox.about(self, "错误", "请先连接到树莓派！")

    # 查看 && 载入已有的签到任务
    def view_button_function(self):
        if self.ISCONNECTED:
            table = TaskTable()
            table.exec_()
            self.task_name = table.task_name
            self.conn.put("/home/xzt/facetest/workbooks/" + self.task_name + ".xlsx", "/home/pi/Project/workbooks/" + self.task_name + ".xlsx")
            with open("now_task_name.txt", 'w', encoding="gbk") as f:
                f.write(self.task_name)
            self.conn.put("/home/xzt/facetest/Window/now_task_name.txt", "/home/pi/Project/now_task_name.txt")
            print(self.task_name)
        else:
            QMessageBox.about(self,"错误","请先连接到树莓派！")

    # 创建faceID
    def create_button_function(self):
        if self.ISCONNECTED:
            id_dialog = OneLineDialog("请输入学号","学号:")
            id_dialog.exec_()
            id_num = id_dialog.text
            #with open('data.txt', 'w') as file:
                #file.writelines(id_num + ' ')
            i_ISCANCEL = id_dialog.CANCEL
            if not i_ISCANCEL:
                name_dialog = OneLineDialog("请输入姓名","姓名:")
                name_dialog.exec_()
                name = name_dialog.text
                #with open('data.txt', 'a') as file:
                    #file.writelines(name)
                n_ISCANCEL = name_dialog.CANCEL
                if not n_ISCANCEL:
                    os.system('mkdir /train/'+ id_num)
                # *********************************************************************** #
                    self.textbrowser.append("Beginning to run the face-ID generator....")
                    #self.conn.put('data.txt','/home/pi/Project/data.txt')
                    Popen("/home/xzt/facetest/exe.sh", stdout=PIPE, close_fds=True)
                    self.timer.timeout.connect(self.print_log)
                    #info = self.conn.exec_command('python3 /home/pi/Project/data_create.py')


        else:
            QMessageBox.about(self,"没有连接","请先连接到树莓派")

    # 查看结果
    def result_button_now_function(self):
        if self.task_name:
            os.system('et /home/xzt/facetest/result/' + self.task_name + '.xlsx')
        else:
            QMessageBox.about(self,"错误", "没有检测到当前任务")

    def result_button_before_function(self):
        QDesktopServices.openUrl(QUrl('file://' + '/home/xzt/facetest/result/'))

    def print_log(self):
        if self.PRINT_LOG:
            self.conn.download("/home/pi/Project/log.txt", "log.txt")
            with open('log.txt', 'r') as f:
                self.log = f.readline()
            if "Finish" not in self.log:
                    self.textbrowser.append(self.log)
                    if "Get" in self.log:
                        self.PRINT_LOG = False
                        self.textbrowser.append("Finish!!! You Can Close the Window Now!!!")
            else:
                pass
        else:
            pass

    def get_now_log(self):
        
        with open("proceed.txt","r") as f:
            c = f.read()
        if self.isTime and eval(c):
            try:
                self.conn.download("/home/pi/Project/now_log.txt", "/home/xzt/facetest/Window/now_log.txt")
                with open("now_log.txt", 'r') as f:
                    self.now_log = f.readline()
                if "Capturing" in self.now_log or self.temp:
                    if self.temp != self.now_log:
                        self.textbrowser.append(self.now_log)
                        self.temp = self.now_log
                self.addtime()
                if not self.b:
                    self.onStart()
                    self.b = True
            except:
                pass
        else:
            pass


if __name__ == "__main__":
    App = QApplication(sys.argv)
    window = Window()
    sys.exit(App.exec())
