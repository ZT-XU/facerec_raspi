from PyQt5.QtWidgets import QWidget,QHBoxLayout,QTableWidget,QPushButton,QApplication,QVBoxLayout,QTableWidgetItem,QCheckBox,QAbstractItemView,QHeaderView,QLabel,QFrame,QMessageBox,QDialog
from PyQt5 import QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont,QColor
import openpyxl
import random
import sys
from Dialog import OneLineDialog


class Table(QDialog):
    def __init__(self):
        super().__init__()
        self.ck_list = []
        self.name_list = []
        self.id_list = []
        self.des_sort = True
        self.final_name_list = []
        self.final_id_list = []
        self.isload = False

        self.initUI()

    def initUI(self):
        self.setWindowTitle('创建任务')
        self.resize(600,480)
        self.table = QTableWidget(self)
        self.btn_create = QPushButton('创建这个任务')
        self.btn_load = QPushButton('载入face-ID')
        self.btn_exit = QPushButton('退出')
        self.spacerItem = QtWidgets.QSpacerItem(20, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.btn_create)
        self.vbox.addWidget(self.btn_load)
        self.vbox.addWidget(self.btn_exit)
        self.vbox.addSpacerItem(self.spacerItem)
        self.txt = QLabel()
        self.txt.setMinimumHeight(50)
        self.vbox2 = QVBoxLayout()
        self.vbox2.addWidget(self.table)
        self.vbox2.addWidget(self.txt)
        self.hbox = QHBoxLayout()
        self.hbox.addLayout(self.vbox2)
        self.hbox.addLayout(self.vbox)
        self.setLayout(self.hbox)
        self.table.setColumnCount(3)
        self.headers = ['选择','学号','姓名']
        self.table.setHorizontalHeaderLabels(self.headers)
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1,200)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setStyleSheet('QHeaderView::section{background:lightblue}')
        self.table.horizontalHeader().setFixedHeight(50)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        #self.table.setColumnHidden(0, True)
        self.table.setStyleSheet('color:green;')
        #self.load_face_id()
        self.btn_load.clicked.connect(self.load_face_id)
        self.btn_create.clicked.connect(self.create_task)
        self.btn_exit.clicked.connect(self.close)
        self.show()

    def load_face_id(self):
        if not self.isload:
            wb = openpyxl.load_workbook("/home/xzt/facetest/FACEID/Face-ID.xlsx")
            ws = wb.active
            count = 0
            for cell in ws['A']:
                value = cell.value
                count += 1
                if value is None:
                    break
                if value == '学号':
                    continue
                self.id_list.append(value)
                self.name_list.append(ws['B' + str(count)].value)
            self.table.setRowCount(count - 1)
            for i in range(count - 1):
                ck = QCheckBox()
                self.table.setItem(i,1,QTableWidgetItem(self.id_list[i]))
                self.table.setItem(i,2,QTableWidgetItem(self.name_list[i]))
                self.table.setCellWidget(i,0,ck)
                self.ck_list.append(ck)
            self.isload = True
        else:
            QMessageBox.about(self,"错误","过多的操作！")

    def create_task(self):
        count = 0
        for ck in self.ck_list:
            if ck.isChecked():
                count += 1
                index = self.ck_list.index(ck)
                self.final_id_list.append(self.id_list[index])
                self.final_name_list.append(self.name_list[index])
        if count == 0:
            QMessageBox.about(self,"错误","没有选择的任务")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "学号"
            ws['B1'] = "姓名"
            ws["C1"] = "是否签到"
            ws["D1"] = "签到时间"
            ws.merge_cells("D1:F1")
            for i in range(count):
                ws['A' + str(i + 2)] = self.final_id_list[i]
                ws["B" + str(i + 2)] = self.final_name_list[i]
            dialog = OneLineDialog("请输入新建任务名", "任务名")
            dialog.exec_()
            task_name = dialog.text
            cancel = dialog.CANCEL
            if not cancel:
                with open('task_name.txt', 'a') as f:
                    f.write(task_name + '\n')
            wb.save("/home/xzt/facetest/workbooks/" + task_name + '.xlsx')
            QMessageBox.about(self,"成功","成功创建任务：" + "\n" + task_name)

class TaskTable(QDialog):
    def __init__(self):
        super().__init__()
        self.task_name_list = []
        self.ck_list = []
        self.task_name = ""
        self.isload = False
        self.final_task_name = []
        self.initUI()

    def initUI(self):
        self.setWindowTitle('加载任务')
        self.resize(460, 480)
        self.table = QTableWidget(self)
        self.btn_load_task = QPushButton('加载已存在的任务')
        self.btn_load = QPushButton('加载这个任务')
        self.btn_exit = QPushButton('退出')
        self.spacerItem = QtWidgets.QSpacerItem(20, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.btn_load_task)
        self.vbox.addWidget(self.btn_load)
        self.vbox.addWidget(self.btn_exit)
        self.vbox.addSpacerItem(self.spacerItem)
        self.txt = QLabel()
        self.txt.setMinimumHeight(50)
        self.vbox2 = QVBoxLayout()
        self.vbox2.addWidget(self.table)
        self.vbox2.addWidget(self.txt)
        self.hbox = QHBoxLayout()
        self.hbox.addLayout(self.vbox2)
        self.hbox.addLayout(self.vbox)
        self.setLayout(self.hbox)
        self.table.setColumnCount(2)
        self.headers = ['选择', '任务']
        self.table.setHorizontalHeaderLabels(self.headers)
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 200)
        self.table.horizontalHeader().setStyleSheet('QHeaderView::section{background:lightblue}')
        self.table.horizontalHeader().setFixedHeight(50)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        # self.table.setColumnHidden(0, True)
        self.table.setStyleSheet('color:green;')
        # self.load_face_id()
        self.btn_load.clicked.connect(self.load_task)
        self.btn_exit.clicked.connect(self.close)
        self.btn_load_task.clicked.connect(self.load_task_name)
        self.show()

    def load_task_name(self):
        if not self.isload:
            with open('/home/xzt/facetest/Window/task_name.txt','r') as f:
                for line in f.readlines():
                    line = line.rstrip()
                    if line:
                        self.task_name_list.append(line)
            count = len(self.task_name_list)
            #print(self.task_name_list)
            self.table.setRowCount(count)
            for i in range(count):
                ck = QCheckBox()
                self.table.setItem(i,1,QTableWidgetItem(self.task_name_list[i]))
                self.table.setCellWidget(i,0,ck)
                self.ck_list.append(ck)
            self.isload = True
        else:
            QMessageBox.about(self,"错误","过多的操作！")

    def load_task(self):
        count = 0
        for ck in self.ck_list:
            if ck.isChecked():
                count += 1
                index = self.ck_list.index(ck)
                self.final_task_name.append(self.task_name_list[index])
        if count > 1:
            QMessageBox.about(self,"错误","选择了过多的任务")
        if count == 1:
                self.task_name = self.final_task_name[0]
                QMessageBox.about(self,"成功","成功地载入了任务：" + "\n" + self.task_name)
        if count == 0:
            QMessageBox.about(self, "错误", "没有选择的任务")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    #table = Table()
    table2 = TaskTable()
    sys.exit(app.exec_())
