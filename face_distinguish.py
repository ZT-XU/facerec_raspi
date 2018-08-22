import face_recognition
import openpyxl
import datetime
import numpy as np 
import picamera
import cv2


camera = picamera.PiCamera()
camera.resolution = (320, 240)
frame = np.empty((240, 320, 3), dtype=n.uint8)

now = datetime.datetime.now()

wb_name = input('Please Enter the Excel File name:')
wb = openpyxl.load_workbook('./workbooks/' + wb_name + '.xlsx')
ws = wb.active
known_face_encodings = []
known_face_names = []
for cells in ws['A']:
    p_name = cells.value
    if p_name == None:
        break
    known_face_names.append(p_name)
    sb_image = face_recognition.load_image_file('./images/' + p_name + '.png')
    known_face_encodings.append(face_recognition.face_encodings(sb_image)[0])

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

while True:
    print("Capturing image ....")
    camera.capture(frame, format="rgb")

    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            # If a match was found in known_face_encodings, just use the first one.
            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]
                ws['B' + str(first_match_index + 1)] = 1
                ws['C' + str(first_match_index + 1)] = now
                print(name + '已签到！' + '    ' + now.strftime('%Y-%m-%d %H:%M:%S'))

            face_names.append(name)

    process_this_frame = not process_this_frame

    cv2.imshow('Video', frame)

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        wb.save('./workbooks/' + wb_name + '.xlsx')
        break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()

