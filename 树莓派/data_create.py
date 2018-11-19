from PIL import Image
import face_recognition
import picamera
import numpy as np
import openpyxl
from time import sleep


with open('/home/pi/Project/data.txt', 'r', encoding="utf-8") as f:
    data_list = f.readline().split(' ')
ID = data_list[0]
Name = data_list[1]


wb_name = 'Face-ID'
wb = openpyxl.load_workbook('/home/pi/Project/FACEID/' + wb_name + '.xlsx')
ws = wb.active
num = 1

value_list = []
for cells in ws['A']:
    value_list.append(cells.value)
    num = num + 1
if ID not in value_list:
    cell = ws.cell(row=num, column=1, value=ID)
    cell = ws.cell(row=num, column=2, value=Name)

camera = picamera.PiCamera()
camera.resolution = (320, 240)
output = np.empty((240, 320, 3), dtype=np.uint8)

while(ID):
    with open('/home/pi/Project/log.txt', 'w') as File:
        File.write("Capturing image...")
    # Grab a single frame of video from the RPi camera as a numpy array
    camera.capture(output, format="rgb")
    face_locations = face_recognition.face_locations(output)

    for face_location in face_locations:

        # Print the location of each face in this image
        top, right, bottom, left = face_location
        

        # You can access the actual face itself like this:
        face_image = output[top-20:bottom+20, left-20:right+20]
        pil_image = Image.fromarray(face_image)
        g_pil_image = pil_image.convert('L')
        g_pil_image.save('/home/pi/Project/images/' + ID + '.png')
        with open('/home/pi/Project/log.txt', 'w') as File:
            File.write("Image Get!!!")
        sleep(2)
        ID = False
wb.save('/home/pi/Project/FACEID/'+ wb_name + '.xlsx')
with open('/home/pi/Project/log.txt','w') as File:
    File.write("Finish!!!")
