import face_recognition
import openpyxl
import datetime
import numpy as np 
#from io import BytesIO
import picamera
import cv2
from time import sleep
from PIL import Image 
import socket
import os,struct
import pickle
from sklearn import neighbors

def predict(img, knn_clf=None, model_path=None, distance_threshold=0.6):
    """
    Recognizes faces in given image using a trained KNN classifier

    :param X_img_path: path to image to be recognized
    :param knn_clf: (optional) a knn classifier object. if not specified, model_save_path must be specified.
    :param model_path: (optional) path to a pickled knn classifier. if not specified, model_save_path must be knn_clf.
    :param distance_threshold: (optional) distance threshold for face classification. the larger it is, the more chance
           of mis-classifying an unknown person as a known one.
    :return: a list of names and face locations for the recognized faces in the image: [(name, bounding box), ...].
        For faces of unrecognized persons, the name 'unknown' will be returned.
    """
    with open(model_path, 'rb') as f:
        knn_clf = pickle.load(f)

    # Load image file and find face locations
    #X_img = face_recognition.load_image_file(img)
    X_face_locations = face_recognition.face_locations(img)

    # If no faces are found in the image, return an empty result.
    if len(X_face_locations) == 0:
        return []

    # Find encodings for faces in the test iamge
    faces_encodings = face_recognition.face_encodings(img, known_face_locations=X_face_locations)

    # Use the KNN model to find the best matches for the test face
    try:
        closest_distances = knn_clf.kneighbors(faces_encodings, n_neighbors=1)
    except:
        return []
    are_matches = [closest_distances[0][i][0] <= distance_threshold for i in range(len(X_face_locations))]
    #print(are_matches)
    return knn_clf.predict(faces_encodings)

host = '192.168.1.100'
port = 5555
sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
address = (host,port)
# 连接服务端
#address_server = ('192.168.1.100',8010)
sock.connect(address)

camera = picamera.PiCamera()
camera.resolution = (320, 240)
frame = np.empty((240, 320, 3), dtype=np.uint8)

now = datetime.datetime.now()

with open("/home/pi/Project/now_task_name.txt",'r',encoding='gbk') as f:
    wb_name = f.read()
wb = openpyxl.load_workbook(u'/home/pi/Project/workbooks/' + wb_name + '.xlsx')
ws = wb.active
known_face_encodings = []
known_face_id_num = []
for cells in ws['A']:
    p_id = cells.value
    if p_id == None:
        break
    if p_id == '学号':
        continue
    known_face_id_num.append(p_id)
    #sb_image = face_recognition.load_image_file('/home/pi/Project/images/' + p_id + '.png')
    #print(p_name,sb_image)
    #known_face_encodings.append(face_recognition.face_encodings(sb_image)[0])
    #bug = face_recognition.face_encodings(sb_image)
    #print(bug)
# Initialize some variables
#frame = BytesIO()
face_locations = []
face_encodings = []
#face_ids = []
print(known_face_id_num)
process_this_frame = True
with open("/home/pi/Project/now_log.txt",'w') as f:
    f.write("Capturing Image....")
while True:
    #print("Capturing image ....")
    #camera.capture(frame, format="rgb")
    camera.capture('/home/pi/Project/frame.jpg')
    filepath = '/home/pi/Project/frame.jpg'
    frame = cv2.imread(filepath)
    if os.path.isfile(filepath):
        fileinfo_size = struct.calcsize('128sl')
        fhead = struct.pack(b'128sq', bytes(os.path.basename(filepath),encoding='utf-8'), os.stat(filepath).st_size)
        #将xxx.jpg以128sq的格式打包
        sock.send(fhead)
        print('client filepath: ', os.path.basename(filepath),os.stat(filepath).st_size)
        fo = open(filepath,'rb')
        while True:
            filedata = fo.read(1024)
            if not filedata:
                break
            sock.send(filedata)
            
    #     fo.close()
    #     print('send over')
    #pil_image = Image.fromarray(frame)

    #pil_image.save("/home/pi/Project/frame.jpg")
    #small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    #rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(frame)
        #face_encodings = face_recognition.face_encodings(frame, face_locations)
        result = predict(frame,model_path='/home/pi/Project/trained_knn_model.clf')
        print(result)
        #face_ids = []
        #for face_encoding in face_encodings:
        #for (top,right,bottom,left),face_encoding in zip(face_locations,face_encodings):
            # See if the face is a match for the known face(s)
            #matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            #name = "Unknown"

            #matches = [i for i,v in enumerate(matches) if v==True]
            # If a match was found in known_face_encodings, just use the first one.
            #for i in matches:
                #first_match_index = i
        if result:
            for name in result:
                if name in known_face_id_num:
                    index = known_face_id_num.index(name)
                    real_name = ws['B' + str(index + 2)].value
                    #name = known_face_id_num[first_match_index]
                    ws['C' + str(index + 2)] = 1
                    ws['D' + str(index + 2)] = now
                    ws.merge_cells('D'+str(index+2)+':F'+str(index+2))
                    with open("/home/pi/Project/now_log.txt",'w') as f:
                        f.writelines( real_name + ' 已签到！！！' + '    ' + now.strftime('%Y-%m-%d %H:%M:%S'))
                    wb.save(u'/home/pi/Project/workbooks/' + wb_name + '.xlsx')

                    #face_names.append(name)
                    #cv2.rectangle(frame,(left,top),(right,bottom),(0,0,255),2)

                    #cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0,255), cv2.FILLED)
                    #pil_image = Image.fromarray(frame)
                    #pil_image.save("/home/pi/Project/frame.jpg")
                else:
                    pass
        with open(filepath,"wb") as img:
            img.write(frame)
                
    process_this_frame = not process_this_frame

    #cv2.imshow('Video', frame)

    # Hit 'q' on the keyboard to quit!
    #if cv2.waitKey(1) & 0xFF == ord('q'):
     #   wb.save('./workbooks/' + wb_name + '.xlsx')
      #  break

# Release handle to the webcam
#video_capture.release()
#cv2.destroyAllWindows()
